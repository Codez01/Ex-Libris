# importers
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment
import pyperclip
import os
#*********************************************************************

abspath = os.path.abspath(__file__)
dname = os.path.dirname(abspath)
os.chdir(dname)

# ********************************************************************
try:
    spam = pyperclip.paste()  # pastes from clipboard
    lTemp = spam.split("=")
    l = lTemp[0].split("\t")

    lreverse = list(reversed(l))

    popCounter = 0

    for i in lreverse:
        if (i == ''):
            popCounter += 1




        else:

            break

    l = list(reversed(lreverse))
    for i in range(0, popCounter):
        l.pop()
    # ********************************************************************
    # ----------counters-----------
    counterList = lTemp[1].replace('\n', '').split("+")
    NA_counter = int(counterList[0])
    EU_counter = int(counterList[1])
    APAC_counter = int(counterList[2])
    ROW_counter = int(counterList[3])
    Global_counter = int(counterList[4])
    # ---------------------------

    d = {}  # dictionary

    # --------------

    total = NA_counter + EU_counter + APAC_counter + ROW_counter + Global_counter
    # ----------------
    l = list(filter(None, l))  # remove empty indexes from a list
    # ********************Looping into the elements *******************************
    j = 0
    for i in range(0, len(l)-7, 8):

        if (l[i + 3] == 'Local'):

            d['local' + str(j)] = l[i + 3]
            d['cloud' + str(j)] = " "
        else:
            d['local' + str(j)] = " "
            d['cloud' + str(j)] = l[i + 3]

        d['case number' + str(j)] = l[i]


        d['customer code' + str(j)] = l[i + 5]


        d['customer name' + str(j)] = l[i + 6]

        d['product' + str(j)] = l[i + 2]

        d['summary' + str(j)] = l[i + 1]

        d['first response' + str(j)] = l[i + 4]

        j = j + 1

    # ************************
    # variables

    STOP_TEXT = "Total"

    NA_row = 8
    EU_row = 0
    APAC_row = 0
    ROW_row = 0
    Global_row = 0
    maxSummaryWidth =[]
    maxNameWidth =[]
    maxCodeWidth =[]

    # -----------------
    NA_one = False
    EU_one = False
    APAC_one = False
    ROW_one = False
    Global_one = False

    # -----------------
    if (NA_counter == 1):
        NA_counter = 0
        NA_one = True

    if (EU_counter == 1):
        EU_counter = 0
        EU_one = True

    if (APAC_counter == 1):
        APAC_counter = 0
        APAC_one = True

    if (ROW_counter == 1):
        ROW_counter = 0
        ROW_one = True

    if (Global_counter == 1):
        Global_counter = 0
        Global_one = True


    # ***************************

    def set_border(ws, cell_range):  # function for adding a border at a specific row
        rows = ws[cell_range]
        side = Side(border_style='thin', color="00000000")  # border type and color

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side

                if pos_x == max_x:
                    border.right = side

                if pos_y == 0:
                    border.top = side
                    border.left = side
                    border.right = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border


    # ************************************

    wb = load_workbook('Empty.xlsx')  # the file we wanna edit
    sheet = wb["Sheet1"]  # select a certain sheet by name

    ws = wb.active
    # ---------------------------------------------------

    # finding a stop

    stopi = 12
    stopj = 4

    # -----------------------Adding ROWS & BORDERS TO IT---------------------

    # ws.insert_rows(8, amount=1)# function for inserting a row at a specific row
    if (NA_counter != 0):
        NA_counter -= 1

        ws.insert_rows(NA_row, amount=NA_counter)  # NA

    EU_row = NA_row + NA_counter + 1

    if (EU_counter != 0):
        EU_counter -= 1

        ws.insert_rows(EU_row, amount=EU_counter)  # eu

    APAC_row = EU_row + EU_counter + 1

    if (APAC_counter != 0):
        APAC_counter -= 1

        ws.insert_rows(APAC_row, amount=APAC_counter)  # apac

    ROW_row = APAC_row + APAC_counter + 1

    if (ROW_counter != 0):
        ROW_counter -= 1

        ws.insert_rows(ROW_row, amount=ROW_counter)  # ROW

    Global_row = ROW_row + ROW_counter + 1

    if (Global_counter != 0):
        Global_counter -= 1
        ws.insert_rows(Global_row, amount=Global_counter)  # Global

    MAX_RANGE = NA_counter + EU_counter + APAC_counter + ROW_counter + Global_counter + stopi  # max range

    for i in range(6, MAX_RANGE):  # loop for the table

        set_border(ws, "D" + str(i) + ":" + "P" + str(i))  # function for adding border for the whole table

    set_border(ws, "D" + str(MAX_RANGE) + ":" + "E" + str(MAX_RANGE))  # function for adding a border at Total row

    # -------------------------------------------------------------------------

    # --------------------adding content to excel file-------------------------------------------

    # add to sheet rows/columns
    if (NA_counter != 0):

        sheet['E7'] = NA_counter + 1
    else:
        sheet['E7'] = NA_counter

    if (EU_counter != 0):

        sheet['E' + str(EU_row - 1)] = EU_counter + 1
    else:
        sheet['E' + str(EU_row - 1)] = EU_counter

    if (APAC_counter != 0):

        sheet['E' + str(APAC_row - 1)] = APAC_counter + 1
    else:
        sheet['E' + str(APAC_row - 1)] = APAC_counter

    if (ROW_counter != 0):

        sheet['E' + str(ROW_row - 1)] = ROW_counter + 1
    else:
        sheet['E' + str(ROW_row - 1)] = ROW_counter

    if (Global_counter != 0):
        sheet['E' + str(Global_row - 1)] = Global_counter + 1
    else:
        sheet['E' + str(Global_row - 1)] = Global_counter

    if (NA_one):
        sheet['E7'] = 1

    if (EU_one):
        sheet['E' + str(EU_row - 1)] = 1

    if (APAC_one):
        sheet['E' + str(APAC_row - 1)] = 1

    if (ROW_one):
        sheet['E' + str(ROW_row - 1)] = 1

    if (Global_one):
        sheet['E' + str(Global_row - 1)] = 1

    # total
    sheet['E' + str(Global_row + Global_counter)] = total

    # ********************************************************************

    NA_counter2 = NA_counter
    EU_counter2 = EU_counter
    APAC_counter2 = APAC_counter
    ROW_counter2 = ROW_counter
    Global_counter2 = Global_counter

    if (NA_one):
        NA_counter = 1
    if (NA_one == False and NA_counter == 0):
        NA_counter = 0
    if (NA_one == False and NA_counter != 0):
        NA_counter += 1

    if (EU_one):
        EU_counter = 1

    if (EU_one == False and EU_counter == 0):
        EU_counter = 0

    if (EU_one == False and EU_counter != 0):
        EU_counter += 1

    if (APAC_one):
        APAC_counter = 1
    if (APAC_one == False and APAC_counter == 0):
        APAC_counter = 0

    if (APAC_one == False and APAC_counter != 0):
        APAC_counter += 1

    if (ROW_one):
        ROW_counter = 1
    if (ROW_one == False and ROW_counter == 0):
        ROW_counter = 0

    if (ROW_one == False and ROW_counter != 0):
        ROW_counter += 1

    if (Global_one):
        Global_counter = 1
    if (Global_one == False and Global_counter == 0):
        Global_counter = 0

    if (Global_one == False and Global_counter != 0):
        Global_counter += 1
    # ********************************************************************
    k = 0  # temp variable

    if (NA_counter != 0):

        for i in range(0, NA_counter, 1):
            sheet["F" + str(7 + i)] = d['local' + str(k)]
            sheet["G" + str(7 + i)] = d['cloud' + str(k)]
            sheet["H" + str(7 + i)] = int(d['case number' + str(k)])
            sheet["I" + str(7 + i)] = d['customer code' + str(k)]
            maxCodeWidth.append(len(d['customer code' + str(k)]))
            sheet["J" + str(7 + i)] = d['customer name' + str(k)]
            maxNameWidth.append(len(d['customer name' + str(k)]))
            sheet["K" + str(7 + i)] = d['product' + str(k)]
            sheet["L" + str(7 + i)] = d['summary' + str(k)]
            maxSummaryWidth.append( len(d['summary' + str(k)]))
            sheet["O" + str(7 + i)] = d['first response' + str(k)] + " min"
            sheet["P"+ str(7+i)] = "No"
            k = k + 1

        k = NA_counter

    if (EU_counter != 0):

        for i in range(EU_row - 1, APAC_row - 1):
            sheet["F" + str(i)] = d['local' + str(k)]
            sheet["G" + str(i)] = d['cloud' + str(k)]
            sheet["H" + str(i)] = int(d['case number' + str(k)])
            sheet["I" + str(i)] = d['customer code' + str(k)]
            maxCodeWidth.append(len(d['customer code' + str(k)]))
            sheet["J" + str(i)] = d['customer name' + str(k)]
            maxNameWidth.append(len(d['customer name' + str(k)]))
            sheet["K" + str(i)] = d['product' + str(k)]
            sheet["L" + str(i)] = d['summary' + str(k)]
            maxSummaryWidth.append(len(d['summary' + str(k)]))
            sheet["O" + str(i)] = d['first response' + str(k)] + " min"
            sheet["P" + str(i)] = "No"
            k = k + 1
        k = NA_counter + EU_counter

    if (APAC_counter != 0):

        for i in range(APAC_row - 1, ROW_row - 1):
            sheet["F" + str(i)] = d['local' + str(k)]
            sheet["G" + str(i)] = d['cloud' + str(k)]
            sheet["H" + str(i)] = int(d['case number' + str(k)])
            sheet["I" + str(i)] = d['customer code' + str(k)]
            maxCodeWidth.append(len(d['customer code' + str(k)]))
            sheet["J" + str(i)] = d['customer name' + str(k)]
            maxNameWidth.append(len(d['customer name' + str(k)]))
            sheet["K" + str(i)] = d['product' + str(k)]
            sheet["L" + str(i)] = d['summary' + str(k)]
            maxSummaryWidth.append(len(d['summary' + str(k)]))
            sheet["O" + str(i)] = d['first response' + str(k)] + " min"
            sheet["P" + str(i)] = "No"
            k = k + 1

        k = NA_counter + EU_counter + APAC_counter

    if (ROW_counter != 0):

        for i in range(ROW_row - 1, Global_row - 1):

            sheet["F" + str(i)] = d['local' + str(k)]
            sheet["G" + str(i)] = d['cloud' + str(k)]

            sheet["H" + str(i)] = int(d['case number' + str(k)])
            sheet["I" + str(i)] = d['customer code' + str(k)]
            maxCodeWidth.append(len(d['customer code' + str(k)]))
            sheet["J" + str(i)] = d['customer name' + str(k)]
            maxNameWidth.append(len(d['customer name' + str(k)]))
            sheet["K" + str(i)] = d['product' + str(k)]
            sheet["L" + str(i)] = d['summary' + str(k)]
            maxSummaryWidth.append(len(d['summary' + str(k)]))
            sheet["O" + str(i)] = d['first response' + str(k)] + " min"
            sheet["P" + str(i)] = "No"
            k = k + 1

        k = NA_counter + EU_counter + APAC_counter + ROW_counter

    if (Global_counter != 0):
        for i in range(Global_row - 1, Global_row + Global_counter - 1):
            sheet["F" + str(i)] = d['local' + str(k)]
            sheet["G" + str(i)] = d['cloud' + str(k)]
            sheet["H" + str(i)] = int(d['case number' + str(k)])
            sheet["I" + str(i)] = d['customer code' + str(k)]
            maxCodeWidth.append(len(d['customer code' + str(k)]))
            sheet["J" + str(i)] = d['customer name' + str(k)]
            maxNameWidth.append(len(d['customer name' + str(k)]))
            sheet["K" + str(i)] = d['product' + str(k)]
            sheet["L" + str(i)] = d['summary' + str(k)]
            maxSummaryWidth.append(len(d['summary' + str(k)]))
            sheet["O" + str(i)] = d['first response' + str(k)] + " min"
            sheet["P" + str(i)] = "No"
            k = k + 1

        k = NA_counter + EU_counter + APAC_counter + ROW_counter + Global_counter
    #***********************************************************************************

    #adjust width of  columns

    sheet.column_dimensions['L'].width =max(maxSummaryWidth)+2

    sheet.column_dimensions['I'].width =max(maxCodeWidth)+4

    sheet.column_dimensions['J'].width = max(maxNameWidth)+2

    # *******************************MERGING CELLS**************************************

    if (NA_counter2 != 0):
        ws.merge_cells("D7" + ":" + "D" + str(7 + NA_counter2))  # merges cells
        ws.merge_cells("E7" + ":" + "E" + str(7 + NA_counter2))  # merges cells

    if (EU_counter2 != 0):
        ws.merge_cells("D" + str(EU_row - 1) + ":" + "D" + str(EU_row + EU_counter2 - 1))  # merges cells
        ws.merge_cells("E" + str(EU_row - 1) + ":" + "E" + str(EU_row + EU_counter2 - 1))  # merges cells

    if (APAC_counter2 != 0):
        ws.merge_cells("D" + str(APAC_row - 1) + ":" + "D" + str(APAC_row + APAC_counter2 - 1))  # merges cells
        ws.merge_cells("E" + str(APAC_row - 1) + ":" + "E" + str(APAC_row + APAC_counter2 - 1))  # merges cells

    if (ROW_counter2 != 0):
        ws.merge_cells("D" + str(ROW_row - 1) + ":" + "D" + str(ROW_row + ROW_counter2 - 1))  # merges cells
        ws.merge_cells("E" + str(ROW_row - 1) + ":" + "E" + str(ROW_row + ROW_counter2 - 1))  # merges cells

    if (Global_counter2 != 0):
        ws.merge_cells("D" + str(Global_row - 1) + ":" + "D" + str(Global_row + Global_counter2 - 1))  # merges cells
        ws.merge_cells("E" + str(Global_row - 1) + ":" + "E" + str(Global_row + Global_counter2 - 1))  # merges cells

    # **********************************
    # centering text
    alignment = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=True,
                          indent=0)

    for i in range(7, MAX_RANGE):

        for j in range(3, 17):
            ws.cell(row=i, column=j).alignment = alignment


    # ------------------------------------------------------------------------------------------

    # end and save changes to file.
    wb.save('DOWNTIMES_YESTERDAY.xlsx')

    print("***********************************************************************************************")
    print("***********************************************************************************************")
    print("**** SUCCESSFULLY COPIED TO EXCEL , OPEN THE FILE BY THE NAME : DOWNTIMES_YESTERDAY.xlsx ******")
    print("**** SUCCESSFULLY COPIED TO EXCEL , OPEN THE FILE BY THE NAME : DOWNTIMES_YESTERDAY.xlsx ******")
    print("**** SUCCESSFULLY COPIED TO EXCEL , OPEN THE FILE BY THE NAME : DOWNTIMES_YESTERDAY.xlsx ******")
    print("***********************************************************************************************")
    print("***********************************************************************************************")
    os.startfile("DOWNTIMES_YESTERDAY.xlsx",'open')# open file automatically

except PermissionError:
    print("**********************************************************************************")
    print("**********************************************************************************")
    print("***** PLEASE CLOSE THE CURRENT OPENED EXCEL FILE AND RUN THE CODE AGAIN!!!! ******")
    print("**********************************************************************************")
    print("**********************************************************************************")

except:
    print("*****************************************************************")
    print("*****************************************************************")
    print("**** MAKE SURE YOU COPIED THE RIGHT RESULT FROM CONSOLE!!! ******")
    print("*****************************************************************")
    print("*****************************************************************")

# *************************************************** END END END END END END END END *********************************************************************
